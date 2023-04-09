import pandas as pd
import openpyxl

tabela = openpyxl.Workbook()

dados = [
    {"User":"koalasmaster","Pass":"koalas"},
    {"User":"paola","Pass":"123456"},
    {"User":"andre","Pass":"210000"}
]
# tabela = pd.DataFrame(dados) #formata em tabela

#Como saber o nome das folhas da tabela
# print(tabela.sheetnames) 

#criar uma folha de tabela nova
# cadastro_db = tabela.create_sheet('Cadastro')

cadastro_db = tabela['Sheet']
cadastro_db.append(['User', 'Pass'])
for i in dados:
    cadastro_db.append([i["User"], i["Pass"]]) #adiciona dados na folha da tabela

tabela.save("Database.xlsx")